import requests_html
import cv2, threading, os
from pyquery import PyQuery as jQuery
import openpyxl, json


class Main:
    # public
    def __init__(this):
        this.session = requests_html.HTMLSession()

    def login(this, user, pwd):
        codeImg = this.__getCodeImg()
        threading.Thread(target=this.__showCodeImg, args=[codeImg]).start()
        data = {"zjh": str(user), "mm": str(pwd), "v_yzm": str(input("验证码："))}
        loginRes = this.session.post('http://urp.hebau.edu.cn/loginAction.do', data=data)
        return True if loginRes.status_code == 200 else False

    def getInfoStructure(this):
        return {
            "name": '',
            "stuid": '',
            "sex": '',
            "id": '',

            "nation": '',
            "nativePlace": '',
            "politicalOutlook": '',
            "birthday": '',

            "clas": '',
            "entrance": '',
            "graduation": '',

            "major": '',

            "Department": '',
            "courses": [],
        }

    def loadInfoById(this, stuid):
        jq = jQuery(this.__getInfoHtml(stuid))
        # info
        tr = jq('table#report1 tr')
        if jq(jq(tr[1]).children('td')[1]).html() is None:
            return None
        stuInfo = {
            "name": jq(jq(tr[1]).children('td')[1]).html(),
            "stuid": jq(jq(tr[1]).children('td')[3]).html(),
            "sex": jq(jq(tr[1]).children('td')[5]).html(),
            "id": jq(jq(tr[1]).children('td')[7]).html(),

            "nation": jq(jq(tr[2]).children('td')[1]).html(),
            "nativePlace": jq(jq(tr[2]).children('td')[3]).html(),
            "politicalOutlook": jq(jq(tr[2]).children('td')[5]).html(),
            "birthday": jq(jq(tr[2]).children('td')[7]).html(),

            "clas": jq(jq(tr[3]).children('td')[1]).html(),
            "entrance": jq(jq(tr[3]).children('td')[3]).html(),
            "graduation": jq(jq(tr[3]).children('td')[5]).html(),

            "major": jq(jq(tr[4]).children('td')[1]).html(),

            "Department": jq(jq(tr[5]).children('td')[1]).html(),
            "courses": [],
        }

        # course
        begin = 7
        end = tr.length - 5
        for i in range(begin, end):
            stuInfo['courses'].append({
                "name": jq(jq(tr[i]).children('td')[0]).html(),
                "credit": jq(jq(tr[i]).children('td')[1]).html(),
                "score": jq(jq(tr[i]).children('td')[2]).html(),
                "method": jq(jq(tr[i]).children('td')[3]).html(),
                "attr": jq(jq(tr[i]).children('td')[4]).html(),
                "time": jq(jq(tr[i]).children('td')[5]).html()
            })
            try:
                if jq(jq(tr[i]).children('td')[6]).html() is not None:
                    stuInfo['courses'].append({
                        "name": jq(jq(tr[i]).children('td')[6]).html(),
                        "credit": jq(jq(tr[i]).children('td')[7]).html(),
                        "score": jq(jq(tr[i]).children('td')[8]).html(),
                        "method": jq(jq(tr[i]).children('td')[9]).html(),
                        "attr": jq(jq(tr[i]).children('td')[10]).html(),
                        "time": jq(jq(tr[i]).children('td')[11]).html()
                    })
            except Exception as e:
                pass
        # end for
        return stuInfo

    def getInfoList(this, idRange):
        infos = []
        for id in idRange:
            info = this.loadInfoById(str(id))
            if info is not None:
                infos.append(info)
                print(info['stuid'])
        return infos

    # private
    def __getInfoHtml(this, stuid):
        data = {"LS_XH": str(stuid), "resultPage": "http://urp.hebau.edu.cn:80/reportFiles/cj/cj_zwcjd.jsp?"}
        res = this.session.post(url="http://urp.hebau.edu.cn/setReportParams", data=data)
        return res.text

    def __showCodeImg(this, data):
        this.__saveTo('./temp.jpg', data)
        codeImg = cv2.imread('./temp.jpg')
        cv2.imshow('code', codeImg)
        cv2.waitKey(0)
        os.remove('./temp.jpg')

    def __getCodeImg(this):
        return this.session.get('http://urp.hebau.edu.cn/validateCodeAction.do?random=0.08322962004793921').content

    def __saveTo(this, dir, data):
        with open(dir, mode='wb') as f:
            f.write(data)


def saveAsJson(dic, outFile):
    jsStr = json.dumps(dic, ensure_ascii=False)
    with open(outFile, 'wt', encoding='utf-8') as f:
        f.write(jsStr)


def get1_5InMem(obj):
    idRange = list(range(2019984040101, 2019984040131)) \
              + list(range(2019984040201, 2019984040231)) \
              + list(range(2019984040301, 2019984040331)) \
              + list(range(2019984040401, 2019984040431)) \
              + list(range(2019984040501, 2019984040531))
    return obj.getInfoList(idRange, './out.json')


# 去除正常考试及补考的信息
def duplicateRemoval(infoList):
    for info in infoList:
        refresher = []
        makeUp = []

        # 仅保留最新一次考试，所以若重修了则需要把补考和正常的删掉，
        for course in info['courses']:
            if course['method'] == '重修':
                refresher.append(course['name'])
        if len(refresher) != 0:
            for course in info['courses']:
                if course['method'] == '正常' and course['name'] in refresher:
                    info['courses'].remove(course)
                if course['method'] == '补考' and course['name'] in refresher:
                    info['courses'].remove(course)

        # 若未重修，但补考了，则需要删掉正常的
        for course in info['courses']:
            if course['method'] == '补考':
                makeUp.append(course['name'])
        if len(makeUp) != 0:
            for course in info['courses']:
                if course['method'] == '正常' and course['name'] in makeUp:
                    info['courses'].remove(course)
    return infoList


def calCreditScore(info):
    scoreSum = 0.0
    creditSum = 0.0
    for course in info['courses']:
        scoreSum += float(course['credit']) * float(course['score'])
        creditSum += float(course['credit'])
    return scoreSum / creditSum


def calCreditScoreDict(infoList):
    creditScoreDict = {}
    for info in infoList:
        creditScore = calCreditScore(info)
        dic = {
            "name": info['name'],
            "stuid": info['stuid'],
            "creditScore": creditScore,
        }
        creditScoreDict[creditScore] = dic
    return creditScoreDict


def sortedCreditScoreList(infoList):
    creditScoreDict = calCreditScoreDict(infoList)
    sortedList = []
    for key in sorted(creditScoreDict.keys(), reverse=True):
        sortedList.append(creditScoreDict[key])
    return sortedList

# 爬出成绩并 save as json
def getAndSaveAsJson(idRange, outSrc, out):
    obj = Main()
    if not (obj.login('****', '****')): exit()
    infoList = obj.getInfoList(idRange)
    saveAsJson(infoList, outSrc)
    infoList = duplicateRemoval(infoList)
    saveAsJson(infoList, out)

# 姓名学号学分成绩和成绩
# scoreTplStuid 是以谁的科目做显示模板的那个人的学号
def saveAsXlsx(infoList, scoreTplStuid, outFile):
    # 为了在后面方便地拿到科目成绩，这里需要散列一下，弄一个 coursesDic
    coursesDic = {}
    for info in infoList:
        courseDic = {}
        for course in info['courses']:
            courseDic[course['name']] = course
        coursesDic[info['stuid']] = courseDic

    # 排序过的 infoList
    sortedList = sortedCreditScoreList(infoList)

    # init excel
    wb = openpyxl.Workbook()
    ws = wb.active

    tplCourses = coursesDic[scoreTplStuid]

    row = 1
    col = 1
    for key in sortedList[0]:
        ws.cell(row=row, column=col).value = key
        col += 1
    courses = []
    for course in tplCourses:
        courses.append(tplCourses[course]['name'])
        ws.cell(row=row, column=col).value = tplCourses[course]['name']
        col += 1
        pass
    row += 1
    for info in sortedList:
        col = 1
        for key in info.keys():
            ws.cell(row=row, column=col).value = info[key]
            col += 1
        for key in courses:
            if key not in coursesDic[info['stuid']]:
                ws.cell(row=row, column=col).value = ''
            else:
                ws.cell(row=row, column=col).value = coursesDic[info['stuid']][key]['score']
            col += 1
        row += 1
        print(info)

    wb.save(outFile)


if __name__ == '__main__':


